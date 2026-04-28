"""Excel dashboard export utilities."""

from __future__ import annotations

from copy import deepcopy
from pathlib import Path

from diamond_gems.config import DATA_OUTPUTS_DIR
from diamond_gems.features.baseline_archetypes import add_baseline_recent_features, assign_pitcher_archetypes, build_what_changed_today_summary

try:
    import pandas as pd  # type: ignore
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
    from openpyxl.styles import Alignment, Font, PatternFill
except ModuleNotFoundError:  # pragma: no cover - project depends on pandas
    pd = None

SHEET_NAME_BY_TABLE = {
    "pitcher_start_summary": "Start Summary",
    "pitcher_pitch_type_summary": "Pitch Type Summary",
    "pitcher_velocity_deltas": "Velocity Deltas",
    "pitcher_usage_deltas": "Usage Deltas",
    "pitcher_trend_scores": "Trend Scores",
    "pitcher_flags": "Flags",
    "content_ideas": "Content Ideas",
}


def _to_records(table) -> list[dict]:
    if pd is not None and isinstance(table, pd.DataFrame):
        return table.to_dict("records")
    if hasattr(table, "to_dict"):
        return table.to_dict("records")
    return deepcopy(table)


def _to_pandas_df(table):
    if pd is None:  # pragma: no cover
        raise ModuleNotFoundError("pandas is required for Excel export.")
    if isinstance(table, pd.DataFrame):
        return table.copy(deep=True)
    return pd.DataFrame.from_records(_to_records(table))


def _first_existing_column(df: pd.DataFrame, options: list[str]) -> str | None:
    for column in options:
        if column in df.columns:
            return column
    return None


def _auto_adjust_column_widths(worksheet) -> None:
    for col_cells in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in col_cells]
        if not values:
            continue
        width = min(max(len(v) for v in values) + 2, 48)
        worksheet.column_dimensions[col_cells[0].column_letter].width = max(width, 10)


def _create_daily_pitcher_board(tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    trend_df = _to_pandas_df(tables.get("pitcher_trend_scores", pd.DataFrame()))
    flags_df = _to_pandas_df(tables.get("pitcher_flags", pd.DataFrame()))
    ideas_df = _to_pandas_df(tables.get("content_ideas", pd.DataFrame()))
    starts_df = _to_pandas_df(tables.get("pitcher_start_summary", pd.DataFrame()))
    velocity_df = _to_pandas_df(tables.get("pitcher_velocity_deltas", pd.DataFrame()))
    usage_df = _to_pandas_df(tables.get("pitcher_usage_deltas", pd.DataFrame()))

    base = trend_df.copy()
    if base.empty and not starts_df.empty:
        base = starts_df.copy()

    pitch_col = _first_existing_column(base, ["pitcher_name"]) or "pitcher_name"
    date_col = _first_existing_column(base, ["game_date", "date"])

    if not flags_df.empty:
        flag_counts = flags_df.groupby("pitcher_name", dropna=False).size().rename("Flag Count").reset_index()
        base = base.merge(flag_counts, on="pitcher_name", how="left")
        top_flag = flags_df.sort_values([c for c in ["confidence_score"] if c in flags_df.columns], ascending=False).drop_duplicates("pitcher_name")
        for src, dst in [("signal", "Top Signal"), ("signal_category", "Signal Category"), ("severity", "Severity"), ("confidence_score", "Confidence")]:
            if src in top_flag.columns:
                base = base.merge(top_flag[["pitcher_name", src]].rename(columns={src: dst}), on="pitcher_name", how="left")
    base["Flag Count"] = pd.to_numeric(base.get("Flag Count", 0), errors="coerce").fillna(0)

    for src_df, src_col, out_col in [
        (velocity_df, "velocity_delta", "Velocity Delta"),
        (usage_df, "usage_delta", "Usage Delta"),
    ]:
        if not src_df.empty and src_col in src_df.columns and "pitcher_name" in src_df.columns:
            sample = src_df.groupby("pitcher_name", dropna=False)[src_col].mean().reset_index()
            base = base.merge(sample.rename(columns={src_col: out_col}), on="pitcher_name", how="left")

    for src, dst in [("whiff_rate", "Whiff %"), ("csw_rate", "CSW %"), ("k_minus_bb_rate", "K-BB %")]:
        if src in base.columns:
            base[dst] = base[src]

    if not ideas_df.empty and "pitcher_name" in ideas_df.columns:
        headline_col = _first_existing_column(ideas_df, ["headline", "angle", "idea"])
        if headline_col is not None:
            top_ideas = ideas_df.drop_duplicates("pitcher_name")[["pitcher_name", headline_col]]
            base = base.merge(top_ideas.rename(columns={headline_col: "Best Content Angle"}), on="pitcher_name", how="left")

    output = pd.DataFrame({
        "Pitcher": base.get(pitch_col, ""),
        "Team": base.get(_first_existing_column(base, ["team_id", "pitcher_team_id"]), ""),
        "Opponent": base.get(_first_existing_column(base, ["opponent_team_id", "opponent"]), ""),
        "Date": base.get(date_col, ""),
        "Overall Trend Score": base.get(_first_existing_column(base, ["overall_trend_score", "pitcher_change_score_percentile", "trend_score"]), ""),
        "Flag Count": base.get("Flag Count", 0),
        "Top Signal": base.get("Top Signal", ""),
        "Signal Category": base.get("Signal Category", ""),
        "Severity": base.get("Severity", ""),
        "Confidence": base.get("Confidence", ""),
        "Velocity Delta": base.get("Velocity Delta", ""),
        "Usage Delta": base.get("Usage Delta", ""),
        "Whiff %": base.get("Whiff %", ""),
        "CSW %": base.get("CSW %", ""),
        "K-BB %": base.get("K-BB %", ""),
        "Best Content Angle": base.get("Best Content Angle", ""),
    })
    output["Overall Trend Score"] = pd.to_numeric(output["Overall Trend Score"], errors="coerce")
    output = output.sort_values(["Overall Trend Score", "Flag Count"], ascending=[False, False], na_position="last").reset_index(drop=True)
    output.insert(0, "Rank", output.index + 1)
    output["Link/Reference to Pitcher Detail"] = output["Pitcher"].astype(str)
    return output


def _create_pitcher_detail_views(tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    pitch_type_df = _to_pandas_df(tables.get("pitcher_pitch_type_summary", pd.DataFrame()))
    if pitch_type_df.empty:
        return pd.DataFrame(columns=["Pitcher", "Date", "Opponent", "Pitch type", "Usage", "Usage delta", "Velocity", "Velocity delta", "Whiff %", "CSW %", "Called strike %", "Chase %", "Zone %", "Notes / generated content angle"])
    rename_map = {
        "pitcher_name": "Pitcher", "game_date": "Date", "opponent_team_id": "Opponent", "pitch_type": "Pitch type",
        "usage_rate": "Usage", "usage_delta": "Usage delta", "avg_velocity": "Velocity", "velocity_delta": "Velocity delta",
        "whiff_rate": "Whiff %", "csw_rate": "CSW %", "called_strike_rate": "Called strike %", "chase_rate": "Chase %", "zone_rate": "Zone %",
    }
    out = pitch_type_df.rename(columns={k: v for k, v in rename_map.items() if k in pitch_type_df.columns}).copy()
    for col in ["Pitcher", "Date", "Opponent", "Pitch type", "Usage", "Usage delta", "Velocity", "Velocity delta", "Whiff %", "CSW %", "Called strike %", "Chase %", "Zone %"]:
        if col not in out.columns:
            out[col] = ""
    out["Notes / generated content angle"] = ""
    return out[["Pitcher", "Date", "Opponent", "Pitch type", "Usage", "Usage delta", "Velocity", "Velocity delta", "Whiff %", "CSW %", "Called strike %", "Chase %", "Zone %", "Notes / generated content angle"]]


def _apply_workbook_styling(workbook) -> None:
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in workbook.worksheets:
        ws.freeze_panes = ws.freeze_panes or "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
        _auto_adjust_column_widths(ws)


def _create_summary_cards(worksheet, board_df: pd.DataFrame) -> int:
    stats = {
        "Total pitchers analyzed": len(board_df),
        "Number of flagged pitchers": int((pd.to_numeric(board_df.get("Flag Count", 0), errors="coerce").fillna(0) > 0).sum()),
        "Average trend score": float(pd.to_numeric(board_df.get("Overall Trend Score"), errors="coerce").mean() or 0),
        "Highest trend score": float(pd.to_numeric(board_df.get("Overall Trend Score"), errors="coerce").max() or 0),
        "Most common signal category": board_df.get("Signal Category", pd.Series(dtype=str)).mode().iloc[0] if not board_df.empty and not board_df.get("Signal Category", pd.Series(dtype=str)).dropna().empty else "",
        "Number of high-severity flags": int((board_df.get("Severity", pd.Series(dtype=str)).astype(str).str.lower() == "high").sum()),
    }
    row = 1
    for label, value in stats.items():
        worksheet.cell(row=row, column=1, value=label).font = Font(bold=True)
        worksheet.cell(row=row, column=2, value=value)
        row += 1
    return row + 1


def _create_charts(worksheet, board_df: pd.DataFrame, first_data_row: int) -> None:
    if board_df.empty:
        return
    max_row = len(board_df) + first_data_row - 1
    if "Overall Trend Score" in board_df.columns:
        chart = BarChart()
        chart.title = "Top 10 by Overall Trend Score"
        data = Reference(worksheet, min_col=6, min_row=first_data_row, max_row=min(max_row, first_data_row + 9))
        cats = Reference(worksheet, min_col=2, min_row=first_data_row, max_row=min(max_row, first_data_row + 9))
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        worksheet.add_chart(chart, "T2")


def export_excel_dashboard(tables: dict[str, "pd.DataFrame"], filename: str = "baseball_content_dashboard.xlsx") -> Path:
    if pd is None:
        raise ModuleNotFoundError("pandas is required for Excel export.")
    output_dir = Path(DATA_OUTPUTS_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / filename

    starts_df = _to_pandas_df(tables.get("pitcher_start_summary", pd.DataFrame()))
    pitch_type_df = _to_pandas_df(tables.get("pitcher_pitch_type_summary", pd.DataFrame()))
    enriched_starts = assign_pitcher_archetypes(add_baseline_recent_features(starts_df, pitch_type_df)) if not starts_df.empty else starts_df
    daily_board_df = _create_daily_pitcher_board({**tables, "pitcher_trend_scores": enriched_starts})
    for col in ["primary_archetype", "secondary_archetypes", "archetype_confidence", "archetype_reason", "baseline_recent_summary", "velo_delta_vs_season", "usage_delta_vs_season", "whiff_delta_vs_season", "csw_delta_vs_season", "kbb_delta_vs_season", "baseline_sample_warning"]:
        if col in enriched_starts.columns and col not in daily_board_df.columns:
            daily_board_df = daily_board_df.merge(enriched_starts[["pitcher_name", "game_date", col]].rename(columns={"pitcher_name":"Pitcher","game_date":"Date"}), on=["Pitcher","Date"], how="left")
    daily_board_df["actionable_today"] = daily_board_df.get("archetype_confidence", "").astype(str).isin(["HIGH", "MEDIUM"])
    detail_df = _create_pitcher_detail_views(tables)
    what_changed_df, what_changed_text = build_what_changed_today_summary(enriched_starts) if not enriched_starts.empty else (pd.DataFrame(), "No strong change signals available today.")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        daily_board_df.to_excel(writer, index=False, sheet_name="Daily Pitcher Board", startrow=8)
        what_changed_df.to_excel(writer, index=False, sheet_name="What Changed Today", startrow=8)
        detail_df.to_excel(writer, index=False, sheet_name="Pitcher Detail Views")
        for table_key, sheet_name in SHEET_NAME_BY_TABLE.items():
            table = tables.get(table_key)
            if table is None:
                continue
            df = _to_pandas_df(table)
            if df.empty:
                continue
            df.to_excel(writer, index=False, sheet_name=sheet_name)

        wb = writer.book
        daily_ws = writer.sheets["Daily Pitcher Board"]
        first_data_row = _create_summary_cards(daily_ws, daily_board_df)
        _create_charts(daily_ws, daily_board_df, first_data_row + 1)
        wc_ws = writer.sheets["What Changed Today"]
        wc_ws.cell(row=1, column=1, value="What Changed Today")
        wc_ws.cell(row=2, column=1, value=what_changed_text)
        _apply_workbook_styling(wb)

        sev_col = 10
        start = first_data_row + 1
        end = daily_ws.max_row
        if end >= start:
            daily_ws.conditional_formatting.add(f"F{start}:F{end}", ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"))
            daily_ws.conditional_formatting.add(f"J{start}:J{end}", FormulaRule(formula=[f'EXACT($J{start},"high")'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))
            daily_ws.conditional_formatting.add(f"L{start}:L{end}", DataBarRule(start_type="num", start_value=-5, end_type="num", end_value=5, color="5B9BD5"))

    return output_path
